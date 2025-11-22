import os
import sys
import shutil
import ctypes
from datetime import date, datetime
from pathlib import Path
from typing import List, Tuple, Optional

from PySide6.QtCore import Qt, QDate, QPoint, QSize, QPointF, QRectF
from PySide6.QtGui import QAction, QIcon, QPainter, QPixmap, QColor, QKeySequence, QFont, QPen
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QDateEdit, QTextEdit, QTableWidget, QTableWidgetItem,
    QAbstractItemView, QHeaderView, QMessageBox, QMenu, QStatusBar, QFrame, QLineEdit,
    QStackedWidget, QCheckBox, QInputDialog, QFileDialog
)

# Excel
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("This tool requires 'openpyxl'. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

APP_TITLE = "Incident Number Tracker"
FILE_NAME = "incident_numbers.xlsx"
SHEET_NAME = "INCIDENTS"
HEADERS = ("Created On", "Ticket ID", "Description")
DATE_NUMBER_FORMAT = "yyyy-mm-dd"

# Activity sheet and datetime formatting
ACTIVITY_SHEET_NAME = "Activity"
ACTIVITY_HEADERS = ("Ticket ID", "Start Time", "End Time")
DATETIME_NUMBER_FORMAT = "yyyy-mm-dd hh:mm:ss"

# UI colours
ACCENT = "#C62828"       # red banner to differentiate
BG_LIGHT = "#FDF7F7"
ROW_ALT = "#FCEAEA"
TEXT_PRIMARY = "#2D1414"


def is_frozen() -> bool:
    return getattr(sys, "frozen", False)


# -------- Helper to find the script/exe directory (for icons) --------

def app_dir() -> Path:
    if is_frozen():
        return Path(sys.executable).resolve().parent
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd()


# -------- Local data directory (Restored to AppData) --------

def get_data_dir() -> Path:
    if sys.platform == "win32":
        base = os.getenv("LOCALAPPDATA")
        if base:
            return Path(base) / "IncidentTracker"
        # Fallback
        return Path.home() / "AppData" / "Local" / "IncidentTracker"
    elif sys.platform == "darwin":
        return Path.home() / "Library" / "Application Support" / "IncidentTracker"
    else:
        return Path.home() / ".local" / "share" / "IncidentTracker"


DATA_DIR = get_data_dir()
DATA_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_PATH = DATA_DIR / FILE_NAME


# -------- OneDrive helper for export --------

def find_onedrive_dir() -> Path:
    # Try environment variables first (Windows)
    for env_name in ("OneDriveCommercial", "OneDriveConsumer", "OneDrive"):
        val = os.getenv(env_name)
        if val and Path(val).exists():
            return Path(val)

    # Common locations
    home = Path.home()
    candidates = [
        home / "OneDrive",
        home / "OneDrive - Personal",
        home / "OneDrive - Microsoft",
        home / "Library" / "CloudStorage",  # macOS OneDrive under CloudStorage
    ]

    # Scan macOS CloudStorage for OneDrive-* folders
    cloud_root = home / "Library" / "CloudStorage"
    if cloud_root.exists():
        for p in cloud_root.iterdir():
            if p.is_dir() and p.name.startswith("OneDrive"):
                candidates.insert(0, p)

    for p in candidates:
        if p.exists():
            return p
    # Fallback to home if nothing found
    return home


# -------- Simple Save Helper (Reverted from Complex Locking) --------

def save_workbook_simple(wb, path: Path):
    """
    Simple save method. Replaces the complex 'atomic save' that caused
    PermissionErrors on Windows.
    """
    try:
        wb.save(path)
    except PermissionError:
        raise PermissionError(
            f"Cannot save the Excel file.\n\n"
            f"The file is currently open in Excel or locked by another process:\n{path}\n\n"
            "Please close the Excel file and try again."
        )


def ensure_workbook_and_sheet(path: Path):
    """
    Ensure workbook exists with required sheets and headers.
    Only saves if created/modified to avoid unnecessary writes.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    dirty = False
    if path.exists():
        try:
            wb = load_workbook(path)
        except Exception:
            # If load fails (corrupt or very locked), try creating new or fail hard
            wb = Workbook()
            dirty = True
    else:
        wb = Workbook()
        dirty = True
        # Remove default 'Sheet' if it's the only one
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb["Sheet"])

    # Ensure INCIDENTS sheet and headers
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(title=SHEET_NAME)
        ws["A1"] = HEADERS[0]
        ws["B1"] = HEADERS[1]
        ws["C1"] = HEADERS[2]
        dirty = True
    else:
        ws = wb[SHEET_NAME]
        if ws["A1"].value is None and ws["B1"].value is None and ws["C1"].value is None:
            ws["A1"] = HEADERS[0]
            ws["B1"] = HEADERS[1]
            ws["C1"] = HEADERS[2]
            dirty = True

    # Ensure Activity sheet and headers
    if ACTIVITY_SHEET_NAME not in wb.sheetnames:
        ws_a = wb.create_sheet(title=ACTIVITY_SHEET_NAME)
        ws_a["A1"] = ACTIVITY_HEADERS[0]
        ws_a["B1"] = ACTIVITY_HEADERS[1]
        ws_a["C1"] = ACTIVITY_HEADERS[2]
        dirty = True
    else:
        ws_a = wb[ACTIVITY_SHEET_NAME]
        if ws_a["A1"].value is None and ws_a["B1"].value is None and ws_a["C1"].value is None:
            ws_a["A1"] = ACTIVITY_HEADERS[0]
            ws_a["B1"] = ACTIVITY_HEADERS[1]
            ws_a["C1"] = ACTIVITY_HEADERS[2]
            dirty = True

    if dirty:
        save_workbook_simple(wb, path)
    return wb


def normalize_description(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    parts = [line.strip() for line in text.split("\n")]
    parts = [p for p in parts if p]
    return ", ".join(parts)


def append_row(path: Path, d: date, ticket_id: str, desc: str):
    wb = ensure_workbook_and_sheet(path)
    ws = wb[SHEET_NAME]
    ws.append([d, ticket_id, desc])
    last_row = ws.max_row
    ws.cell(row=last_row, column=1).number_format = DATE_NUMBER_FORMAT
    save_workbook_simple(wb, path)


def read_rows(path: Path) -> List[Tuple[str, str, str]]:
    if not path.exists():
        return []
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return []
        
    if SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[SHEET_NAME]
    rows: List[Tuple[str, str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        created_on, ticket_id, description = (row + (None, None, None))[:3]
        if created_on is None and ticket_id is None and description is None:
            continue
        if isinstance(created_on, (datetime, date)):
            dstr = created_on.strftime("%Y-%m-%d")
        elif created_on:
            dstr = str(created_on)
        else:
            dstr = ""
        rows.append((dstr, str(ticket_id or ""), description or ""))
    return rows


def next_default_ticket_for_date(path: Path, d: date) -> str:
    """
    Compute next 'THyymmddNN' for the given date by scanning existing Ticket IDs.
    Only considers IDs that start with the THyymmdd prefix.
    """
    yy = d.year % 100
    prefix = f"TH{yy:02d}{d.month:02d}{d.day:02d}"
    max_seq = 0
    try:
        if path.exists():
            wb = load_workbook(path, data_only=True)
            if SHEET_NAME in wb.sheetnames:
                ws = wb[SHEET_NAME]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    ticket_id = (row[1] or "").strip()
                    if ticket_id.startswith(prefix):
                        tail = ticket_id[len(prefix):]
                        if tail.isdigit():
                            max_seq = max(max_seq, int(tail))
    except Exception:
        pass
    return f"{prefix}{max_seq + 1:02d}"


def emoji_icon(emoji: str, size: int = 128,
               bg=QColor(198, 40, 40), fg=QColor(255, 255, 255)) -> QIcon:
    pm = QPixmap(size, size)
    pm.fill(Qt.transparent)
    painter = QPainter(pm)
    painter.setRenderHint(QPainter.Antialiasing, True)
    painter.setBrush(bg)
    painter.setPen(Qt.NoPen)
    painter.drawEllipse(0, 0, size, size)
    painter.setPen(fg)
    painter.drawText(pm.rect(), Qt.AlignCenter, emoji)
    painter.end()
    return QIcon(pm)


def load_window_icon() -> QIcon:
    # Try common icon names in the application directory
    for name in ("app.ico", "app.png", "app.icns"):
        p = app_dir() / name
        if p.exists():
            return QIcon(str(p))
    # Fallback to the emoji-generated icon if not found
    return emoji_icon("🚨")


def sort_az_pixmap(size: int = 20, fg=QColor(45, 20, 20)) -> QPixmap:
    pm = QPixmap(size, size)
    pm.fill(Qt.transparent)
    p = QPainter(pm)
    p.setRenderHint(QPainter.Antialiasing, True)

    pen = QPen(fg)
    pen.setWidth(max(1, int(size * 0.08)))
    p.setPen(pen)

    # Down arrow on the left
    m = size * 0.18
    x = m + size * 0.08
    y_top = m
    y_bot = size - m - size * 0.1
    p.drawLine(QPointF(x, y_top), QPointF(x, y_bot))
    ah = size * 0.22
    p.drawLine(QPointF(x, y_bot), QPointF(x - ah, y_bot - ah))
    p.drawLine(QPointF(x, y_bot), QPointF(x + ah, y_bot - ah))

    # A and Z on the right
    f = QFont()
    f.setBold(True)
    f.setPointSizeF(size * 0.52)
    p.setFont(f)

    right_x = size * 0.46
    half_h = size * 0.48
    p.drawText(QRectF(right_x, 0, size - right_x, half_h), Qt.AlignLeft | Qt.AlignVCenter, "A")
    p.drawText(QRectF(right_x, half_h - size * 0.06, size - right_x, half_h), Qt.AlignLeft | Qt.AlignVCenter, "Z")

    p.end()
    return pm


# -------- Activity helpers --------

def append_activity_start(path: Path, ticket_id: str, start_dt: datetime):
    wb = ensure_workbook_and_sheet(path)
    ws = wb[ACTIVITY_SHEET_NAME]
    ws.append([ticket_id, start_dt, None])
    last_row = ws.max_row
    ws.cell(row=last_row, column=2).number_format = DATETIME_NUMBER_FORMAT
    save_workbook_simple(wb, path)


def find_latest_open_activity_row(ws, ticket_id: str) -> int:
    # Returns row index (>1) of latest row for ticket_id with empty End Time, else 0
    for row_idx in range(ws.max_row, 1, -1):
        a = ws.cell(row=row_idx, column=1).value
        c = ws.cell(row=row_idx, column=3).value
        a_str = str(a or "").strip()
        c_empty = c is None or (isinstance(c, str) and not c.strip())
        if a_str == ticket_id and c_empty:
            return row_idx
    return 0


def set_activity_end(path: Path, ticket_id: str, end_dt: datetime) -> bool:
    wb = ensure_workbook_and_sheet(path)
    ws = wb[ACTIVITY_SHEET_NAME]
    row_idx = find_latest_open_activity_row(ws, ticket_id)
    if row_idx == 0:
        return False
    ws.cell(row=row_idx, column=3, value=end_dt)
    ws.cell(row=row_idx, column=3).number_format = DATETIME_NUMBER_FORMAT
    save_workbook_simple(wb, path)
    return True


def has_open_activity(path: Path, ticket_id: str) -> bool:
    if not path.exists():
        return False
    wb = load_workbook(path)
    if ACTIVITY_SHEET_NAME not in wb.sheetnames:
        return False
    ws = wb[ACTIVITY_SHEET_NAME]
    return find_latest_open_activity_row(ws, ticket_id) > 0


def read_activity_rows_for_ticket(path: Path, ticket_id: str) -> List[Tuple[str, str, str]]:
    """Return list of (Ticket ID, Start Time str, End Time str) for a given ticket, sorted by Start Time."""
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    if ACTIVITY_SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[ACTIVITY_SHEET_NAME]
    rows: List[Tuple[datetime, str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t_id, start, end = (row + (None, None, None))[:3]
        if str(t_id or "").strip() != ticket_id:
            continue

        def fmt_dt(x):
            if isinstance(x, datetime):
                return x.strftime("%Y-%m-%d %H:%M:%S")
            return str(x or "")
        rows.append((start if isinstance(start, datetime) else datetime.min, fmt_dt(start), fmt_dt(end)))
    rows.sort(key=lambda r: r[0])
    return [(ticket_id, r[1], r[2]) for r in rows]


# -------- New helpers for updating Ticket IDs --------

def update_ticket_id_in_incidents_by_row(path: Path, excel_row_idx: int, new_ticket_id: str):
    wb = ensure_workbook_and_sheet(path)
    ws = wb[SHEET_NAME]
    ws.cell(row=excel_row_idx, column=2, value=new_ticket_id)  # column 2 = Ticket ID
    save_workbook_simple(wb, path)


def rename_ticket_id_in_activity(path: Path, old_id: str, new_id: str) -> int:
    wb = ensure_workbook_and_sheet(path)
    ws = wb[ACTIVITY_SHEET_NAME]
    changed = 0
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if str(a or "").strip() == old_id:
            ws.cell(row=r, column=1, value=new_id)
            changed += 1
    save_workbook_simple(wb, path)
    return changed


def any_activity_for_ticket(path: Path, ticket_id: str) -> bool:
    if not path.exists():
        return False
    wb = load_workbook(path, data_only=True)
    if ACTIVITY_SHEET_NAME not in wb.sheetnames:
        return False
    ws = wb[ACTIVITY_SHEET_NAME]
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if str(a or "").strip() == ticket_id:
            return True
    return False


def ticket_id_exists_elsewhere(path: Path, new_id: str, exclude_row: int) -> bool:
    if not path.exists():
        return False
    wb = load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        return False
    ws = wb[SHEET_NAME]
    for r in range(2, ws.max_row + 1):
        if r == exclude_row:
            continue
        val = ws.cell(row=r, column=2).value
        if str(val or "").strip() == new_id:
            return True
    return False


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.setMinimumSize(980, 700)
        self.setWindowIcon(load_window_icon())  # use app.ico if present

        ensure_workbook_and_sheet(EXCEL_PATH)

        self._user_edited_ticket = False  # track if user has typed in Ticket ID
        self._activity_ticket_id: str = ""  # current ticket shown in Activity view
        self.latest_first: bool = False  # user-controlled sort state

        # Central with stacked views
        central = QWidget()
        central_layout = QVBoxLayout(central)
        central_layout.setContentsMargins(0, 0, 0, 0)
        central_layout.setSpacing(0)
        self.stack = QStackedWidget()
        central_layout.addWidget(self.stack)
        self.setCentralWidget(central)

        # Build pages
        self.page_main = QWidget()
        self.page_activity = QWidget()
        self._build_main_page(self.page_main)
        self._build_activity_page(self.page_activity)

        self.stack.addWidget(self.page_main)
        self.stack.addWidget(self.page_activity)
        self.stack.setCurrentWidget(self.page_main)

        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.update_status(f"File: {EXCEL_PATH}")

        # Shortcuts
        self._add_shortcuts()

        # Initial fill and load
        self.fill_default_ticket()
        self.load_table()
        self.update_preview()

        # Styling
        self._apply_styles()

    def _build_main_page(self, page: QWidget):
        root = QVBoxLayout(page)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        # Banner
        banner = QFrame()
        banner.setStyleSheet(f"background:{ACCENT}; border-radius:6px;")
        bl = QVBoxLayout(banner)
        title = QLabel(APP_TITLE)
        subtitle = QLabel("Track incident tickets quickly (with smart default IDs). Double-click a ticket to manage Activity.")
        title.setStyleSheet("color:white; font-size:18px; font-weight:600;")
        subtitle.setStyleSheet("color:#FFE5E5; font-size:12px;")
        bl.addWidget(title)
        bl.addWidget(subtitle)
        root.addWidget(banner)

        # Top row: date + ticket id
        top = QHBoxLayout()

        lbl_date = QLabel("Created On:")
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.dateChanged.connect(self.on_date_changed)
        btn_today = QPushButton("Today")
        btn_today.clicked.connect(self.set_today)

        top.addWidget(lbl_date)
        top.addWidget(self.date_edit)
        top.addWidget(btn_today)
        top.addSpacing(20)

        lbl_ticket = QLabel("Ticket ID:")
        self.ticket_edit = QLineEdit()
        self.ticket_edit.setPlaceholderText("e.g. INC2453903 or leave default THyymmddNN")
        self.ticket_edit.textEdited.connect(self.on_ticket_edited)
        btn_fill_default = QPushButton("Fill Default")
        btn_fill_default.clicked.connect(self.fill_default_ticket)

        top.addWidget(lbl_ticket)
        top.addWidget(self.ticket_edit)
        top.addWidget(btn_fill_default)
        top.addStretch(1)
        root.addLayout(top)

        # Description
        desc_frame = QVBoxLayout()
        lbl_desc = QLabel("Description (multi-line allowed)")
        self.desc_text = QTextEdit()
        self.desc_text.setPlaceholderText("Enter incident description; multiple lines will be joined with commas")
        self.desc_text.textChanged.connect(self.update_preview)
        desc_frame.addWidget(lbl_desc)
        desc_frame.addWidget(self.desc_text)
        root.addLayout(desc_frame)

        # Preview
        lbl_prev = QLabel("Preview (single line):")
        self.preview = QLabel("(nothing yet)")
        self.preview.setStyleSheet("background:white; border:1px solid #E6BFBF; padding:8px;")
        root.addWidget(lbl_prev)
        root.addWidget(self.preview)

        # Buttons
        btn_row = QHBoxLayout()
        self.btn_add = QPushButton("Add")
        self.btn_add.clicked.connect(self.on_add)

        # New: Update Ticket ID button
        self.btn_update = QPushButton("Update Ticket ID")
        self.btn_update.setToolTip("Update the selected row's Ticket ID (and optionally Activity)")
        self.btn_update.clicked.connect(self.on_update_ticket_id)

        self.btn_clear = QPushButton("Clear")
        self.btn_clear.clicked.connect(self.on_clear)
        self.btn_open = QPushButton("Open Excel")
        self.btn_open.clicked.connect(self.open_excel)

        # New: Export to OneDrive
        self.btn_export = QPushButton("Export to OneDrive…")
        self.btn_export.setToolTip("Copy the current workbook to your OneDrive folder")
        self.btn_export.clicked.connect(self.on_export_to_onedrive)

        self.btn_refresh = QPushButton("Refresh List")
        self.btn_refresh.clicked.connect(self.load_table)

        # Strongly-visible checkbox + clickable A↓Z icon label
        self.chk_latest_first = QCheckBox()
        self.chk_latest_first.setTristate(False)
        self.chk_latest_first.setToolTip("Latest First")
        self.chk_latest_first.setStyleSheet("""
        QCheckBox::indicator {
          width: 20px; height: 20px;
          border: 2px solid #8A5555;
          border-radius: 4px;
          background: #FFFFFF;
        }
        QCheckBox::indicator:hover { border-color: #C62828; }
        QCheckBox::indicator:checked {
          background: #2E7D32;
          border-color: #2E7D32;
        }
        """)
        self.chk_latest_first.toggled.connect(self.on_toggle_latest_first)

        self.lbl_latest_icon = QLabel()
        self.lbl_latest_icon.setToolTip("Latest First")
        self.lbl_latest_icon.setCursor(Qt.PointingHandCursor)

        def _toggle_checkbox(_event):
            self.chk_latest_first.toggle()
        self.lbl_latest_icon.mousePressEvent = _toggle_checkbox  # type: ignore

        self.btn_view_activity = QPushButton("View Activity")
        self.btn_view_activity.setToolTip("Open Activity view for the selected Ticket ID")
        self.btn_view_activity.clicked.connect(self.open_activity_for_selection)

        tip = QLabel("Shortcuts: Ctrl+Enter Add • Ctrl+L Clear • Ctrl+T Today • Ctrl+O Open • F5 Refresh • Double-click a row to view Activity")
        tip.setStyleSheet("color:#8A5555;")

        btn_row.addWidget(self.btn_add)
        btn_row.addWidget(self.btn_update)
        btn_row.addWidget(self.btn_clear)
        btn_row.addWidget(self.btn_open)
        btn_row.addWidget(self.btn_export)
        btn_row.addWidget(self.btn_refresh)

        # Add checkbox and its icon label
        btn_row.addWidget(self.chk_latest_first)
        btn_row.addWidget(self.lbl_latest_icon)

        btn_row.addWidget(self.btn_view_activity)
        btn_row.addStretch(1)
        btn_row.addWidget(tip)
        root.addLayout(btn_row)

        # Table (3 columns)
        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(HEADERS)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.on_table_context_menu)
        self.table.itemDoubleClicked.connect(self.on_table_double_clicked)
        root.addWidget(self.table)

        # Set initial icon tint based on current state
        self._update_latest_icon(self.chk_latest_first.isChecked())

    def _build_activity_page(self, page: QWidget):
        root = QVBoxLayout(page)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        # Header with Back and title
        header = QHBoxLayout()
        self.btn_back = QPushButton("← Back")
        self.btn_back.clicked.connect(self.back_to_main)
        self.act_title = QLabel("Activity")
        self.act_title.setStyleSheet("font-size:16px; font-weight:600;")
        header.addWidget(self.btn_back)
        header.addSpacing(8)
        header.addWidget(self.act_title)
        header.addStretch(1)

        # Controls on right: Start/Stop/Refresh
        self.act_btn_start = QPushButton("Start")
        self.act_btn_start.setToolTip("Start activity for this Ticket ID (date & time)")
        self.act_btn_start.clicked.connect(self.on_start_activity)
        self.act_btn_stop = QPushButton("Stop")
        self.act_btn_stop.setToolTip("Stop latest open activity for this Ticket ID")
        self.act_btn_stop.clicked.connect(self.on_stop_activity)
        self.act_btn_refresh = QPushButton("Refresh")
        self.act_btn_refresh.clicked.connect(self.load_activity_table)

        header.addWidget(self.act_btn_refresh)
        header.addWidget(self.act_btn_start)
        header.addWidget(self.act_btn_stop)
        root.addLayout(header)

        # Info
        self.act_info = QLabel("")
        self.act_info.setStyleSheet("color:#8A5555;")
        root.addWidget(self.act_info)

        # Activity table
        self.act_table = QTableWidget(0, 3)
        self.act_table.setHorizontalHeaderLabels(ACTIVITY_HEADERS)  # Ticket ID, Start Time, End Time
        self.act_table.setAlternatingRowColors(True)
        self.act_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.act_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.act_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.act_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.act_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        root.addWidget(self.act_table)

        tip = QLabel("Tip: Use Ctrl+Shift+S to Start and Ctrl+E to Stop while on this view.")
        tip.setStyleSheet("color:#8A5555;")
        root.addWidget(tip)

    def _apply_styles(self):
        self.setStyleSheet(f"""
            QWidget {{ background: {BG_LIGHT}; color: {TEXT_PRIMARY}; }}
            QTextEdit {{ background: white; border: 1px solid #E6BFBF; }}
            QTableWidget {{ background: white; alternate-background-color: {ROW_ALT}; }}
            QPushButton {{ padding: 6px 10px; }}
        """)

    def _add_shortcuts(self):
        def add_seq(seq, handler):
            act = QAction(self)
            act.setShortcut(QKeySequence(seq))
            act.triggered.connect(handler)
            self.addAction(act)

        add_seq("Ctrl+Return", self.on_add)
        add_seq("Ctrl+Enter", self.on_add)
        add_seq("Ctrl+L", self.on_clear)
        add_seq("Ctrl+T", self.set_today)
        add_seq("Ctrl+O", self.open_excel)
        add_seq("F5", self.load_table)
        # Activity view shortcuts
        add_seq("Ctrl+Shift+S", self.on_start_activity)
        add_seq("Ctrl+E", self.on_stop_activity)

    # -------- Helpers --------

    def _update_latest_icon(self, checked: bool):
        # Green when ON, grey when OFF
        colour = QColor(46, 125, 50) if checked else QColor(122, 122, 122)
        self.lbl_latest_icon.setPixmap(sort_az_pixmap(20, fg=colour))

    def _select_row_by_excel_row(self, excel_row_idx: int):
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0)
            if item and item.data(Qt.UserRole) == excel_row_idx:
                self.table.selectRow(r)
                self.table.scrollToItem(item, QAbstractItemView.PositionAtCenter)
                break

    # -------- Main view handlers --------

    def set_today(self):
        self.date_edit.setDate(QDate.currentDate())

    def on_date_changed(self):
        # Only auto-update if the user hasn't started typing a manual ID
        if not self._user_edited_ticket:
            self.fill_default_ticket()

    def on_ticket_edited(self):
        self._user_edited_ticket = True

    def fill_default_ticket(self):
        dqt = self.date_edit.date()
        d_py = date(dqt.year(), dqt.month(), dqt.day())
        suggested = next_default_ticket_for_date(EXCEL_PATH, d_py)
        self.ticket_edit.setText(suggested)
        self._user_edited_ticket = False  # treat the current value as app-suggested

    def on_clear(self):
        self.desc_text.clear()
        self.update_preview()
        # Keep date and ticket; often you’ll add multiple entries

    def on_add(self):
        # Date
        dqt = self.date_edit.date()
        d_py = date(dqt.year(), dqt.month(), dqt.day())

        # Ticket ID
        ticket_id = self.ticket_edit.text().strip()
        if not ticket_id:
            ticket_id = next_default_ticket_for_date(EXCEL_PATH, d_py)

        desc = normalize_description(self.desc_text.toPlainText())
        if not desc:
            QMessageBox.critical(self, "Missing description", "Please enter the Description.")
            return

        try:
            append_row(EXCEL_PATH, d_py, ticket_id, desc)
        except PermissionError as e:
            QMessageBox.critical(self, "Cannot save", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add entry:\n{e}")
            return

        self.update_status(f"Added: {d_py.strftime('%Y-%m-%d')} • {ticket_id}")

        # Prepare for next entry:
        self.on_clear()
        self.fill_default_ticket()
        self.load_table()

    def update_preview(self):
        text = self.desc_text.toPlainText().strip()
        combined = normalize_description(text)
        self.preview.setText(combined or "(nothing yet)")

    def on_toggle_latest_first(self, checked: bool):
        # Update icon tint
        self._update_latest_icon(checked)
        # Sorting behaviour
        self.latest_first = checked
        if checked:
            self.table.setSortingEnabled(True)
            self.table.sortByColumn(0, Qt.DescendingOrder)  # Created On column
        else:
            # Disable sorting and reload to restore the original order from the file
            self.table.setSortingEnabled(False)
            self.load_table()

    def load_table(self):
        try:
            ensure_workbook_and_sheet(EXCEL_PATH)
            wb = load_workbook(EXCEL_PATH, data_only=True)
            if SHEET_NAME not in wb.sheetnames:
                rows = []
            else:
                ws = wb[SHEET_NAME]
                rows = []
                for excel_row in range(2, ws.max_row + 1):
                    created_on = ws.cell(row=excel_row, column=1).value
                    ticket_id = ws.cell(row=excel_row, column=2).value
                    description = ws.cell(row=excel_row, column=3).value
                    # Skip completely empty rows
                    if created_on is None and ticket_id is None and description is None:
                        continue
                    if isinstance(created_on, (datetime, date)):
                        dstr = created_on.strftime("%Y-%m-%d")
                    elif created_on:
                        dstr = str(created_on)
                    else:
                        dstr = ""
                    rows.append((excel_row, dstr, str(ticket_id or ""), description or ""))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read Excel:\n{e}")
            return

        # Disable sorting while populating to avoid reordering during insertions
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        for excel_row, dstr, ticket, desc in rows:
            r = self.table.rowCount()
            self.table.insertRow(r)
            it0 = QTableWidgetItem(dstr)
            it1 = QTableWidgetItem(ticket)
            it2 = QTableWidgetItem(desc)
            # Store the Excel row index on the items (UserRole)
            it0.setData(Qt.UserRole, excel_row)
            it1.setData(Qt.UserRole, excel_row)
            it2.setData(Qt.UserRole, excel_row)
            self.table.setItem(r, 0, it0)
            self.table.setItem(r, 1, it1)
            self.table.setItem(r, 2, it2)

        # Re-apply sorting if user wants latest first
        if self.latest_first:
            self.table.setSortingEnabled(True)
            self.table.sortByColumn(0, Qt.DescendingOrder)

        self.update_status(f"File: {EXCEL_PATH} • records: {len(rows)}")

    def on_table_context_menu(self, pos: QPoint):
        idx = self.table.indexAt(pos)
        if not idx.isValid():
            return
        self.table.selectRow(idx.row())
        menu = QMenu(self)
        act_copy = QAction("Copy row", self)
        act_copy.triggered.connect(self.copy_selected_row)
        act_open = QAction("View Activity", self)
        act_open.triggered.connect(self.open_activity_for_selection)
        act_update = QAction("Update Ticket ID...", self)
        act_update.triggered.connect(self.on_update_ticket_id)
        menu.addAction(act_copy)
        menu.addAction(act_open)
        menu.addSeparator()
        menu.addAction(act_update)
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def copy_selected_row(self):
        row = self.table.currentRow()
        if row < 0:
            return
        d = self.table.item(row, 0).text() if self.table.item(row, 0) else ""
        ticket = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
        desc = self.table.item(row, 2).text() if self.table.item(row, 2) else ""
        text = f"{d}\t{ticket}\t{desc}"
        QApplication.clipboard().setText(text)
        self.update_status("Row copied to clipboard.")

    def on_table_double_clicked(self, item: QTableWidgetItem):
        self.open_activity_for_selection()

    def open_activity_for_selection(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "No selection", "Please select a ticket row first.")
            return
        item = self.table.item(row, 1)
        if not item or not item.text().strip():
            QMessageBox.information(self, "No Ticket ID", "Selected row has no Ticket ID.")
            return
        ticket_id = item.text().strip()
        self.open_activity_view_for_ticket(ticket_id)

    # -------- Update Ticket ID handler --------

    def on_update_ticket_id(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "No selection", "Please select a row to update.")
            return

        item_ticket = self.table.item(row, 1)
        item_any = self.table.item(row, 0)
        if not item_ticket or not item_any:
            QMessageBox.information(self, "Invalid selection", "Selected row is invalid.")
            return

        old_ticket = item_ticket.text().strip()
        excel_row_idx = item_any.data(Qt.UserRole)
        if not isinstance(excel_row_idx, int):
            QMessageBox.critical(self, "Error", "Internal reference to Excel row not found.")
            return

        # Default proposed value: current Ticket Edit if non-empty, otherwise the old ticket
        default_text = self.ticket_edit.text().strip() or old_ticket

        new_ticket, ok = QInputDialog.getText(
            self, "Update Ticket ID",
            f"Enter new Ticket ID for the selected row:\nOld: {old_ticket}",
            text=default_text
        )
        if not ok:
            return
        new_ticket = new_ticket.strip()
        if not new_ticket:
            QMessageBox.information(self, "No change", "New Ticket ID cannot be empty.")
            return
        if new_ticket == old_ticket:
            self.update_status("Ticket ID unchanged.")
            return

        # Optional duplicate check
        try:
            if ticket_id_exists_elsewhere(EXCEL_PATH, new_ticket, exclude_row=excel_row_idx):
                resp = QMessageBox.question(
                    self, "Duplicate Ticket ID",
                    f"The Ticket ID '{new_ticket}' already exists in the list.\n"
                    "Do you still want to proceed?",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )
                if resp != QMessageBox.Yes:
                    return
        except Exception:
            pass  # non-fatal; allow update

        # Update INCIDENTS
        try:
            update_ticket_id_in_incidents_by_row(EXCEL_PATH, excel_row_idx, new_ticket)
        except PermissionError as e:
            QMessageBox.critical(self, "Cannot save", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to update Ticket ID:\n{e}")
            return

        # Optionally update Activity sheet if there are records for old_ticket
        changed = 0
        try:
            if any_activity_for_ticket(EXCEL_PATH, old_ticket):
                resp = QMessageBox.question(
                    self, "Also update Activity?",
                    f"Activity entries exist for '{old_ticket}'.\n"
                    f"Do you want to rename them to '{new_ticket}' as well?",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes
                )
                if resp == QMessageBox.Yes:
                    changed = rename_ticket_id_in_activity(EXCEL_PATH, old_ticket, new_ticket)
                    self.update_status(f"Ticket ID updated; Activity entries renamed: {changed}")
                else:
                    self.update_status("Ticket ID updated (Activity not changed).")
            else:
                self.update_status("Ticket ID updated.")
        except PermissionError as e:
            QMessageBox.critical(self, "Cannot save", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to update Activity:\n{e}")
            return

        # Refresh and reselect the updated row
        self.load_table()
        self._select_row_by_excel_row(excel_row_idx)

        # If Activity view is open for the old ticket and we renamed, update the view
        if changed > 0 and self._activity_ticket_id == old_ticket:
            self._activity_ticket_id = new_ticket
            self.act_title.setText(f"Activity: {new_ticket}")
            self.load_activity_table()

        # Update the input field to reflect new ticket
        self.ticket_edit.setText(new_ticket)

    # -------- Activity view handlers --------

    def current_ticket_from_selection_or_edit(self) -> str:
        # Prefer the Activity view's current ticket if set; otherwise selected row; fallback to ticket_edit
        if self._activity_ticket_id:
            return self._activity_ticket_id
        row = self.table.currentRow()
        if row >= 0:
            item = self.table.item(row, 1)  # column 1 is Ticket ID
            if item and item.text().strip():
                return item.text().strip()
        return self.ticket_edit.text().strip()

    def open_activity_view_for_ticket(self, ticket_id: str):
        self._activity_ticket_id = ticket_id
        self.act_title.setText(f"Activity: {ticket_id}")
        self.act_info.setText("")
        self.load_activity_table()
        self.stack.setCurrentWidget(self.page_activity)
        self.update_status(f"Viewing Activity for {ticket_id}")

    def back_to_main(self):
        self._activity_ticket_id = ""
        self.stack.setCurrentWidget(self.page_main)
        self.update_status(f"File: {EXCEL_PATH}")

    def load_activity_table(self):
        ticket_id = self._activity_ticket_id
        if not ticket_id:
            self.act_info.setText("No Ticket ID selected.")
            self.act_table.setRowCount(0)
            return

        try:
            rows = read_activity_rows_for_ticket(EXCEL_PATH, ticket_id)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read Activity:\n{e}")
            return

        self.act_table.setRowCount(0)
        for t, start_str, end_str in rows:
            r = self.act_table.rowCount()
            self.act_table.insertRow(r)
            self.act_table.setItem(r, 0, QTableWidgetItem(t))
            self.act_table.setItem(r, 1, QTableWidgetItem(start_str))
            self.act_table.setItem(r, 2, QTableWidgetItem(end_str))

        # Show open/closed status
        try:
            open_running = has_open_activity(EXCEL_PATH, ticket_id)
        except Exception:
            open_running = False
        self.act_info.setText("Status: Running" if open_running else "Status: Stopped")

    def on_start_activity(self):
        ticket_id = self.current_ticket_from_selection_or_edit()
        if not ticket_id:
            QMessageBox.warning(self, "No Ticket ID", "Select or open a Ticket ID before starting.")
            return

        # Prevent overlapping open sessions for the same ticket
        try:
            if has_open_activity(EXCEL_PATH, ticket_id):
                QMessageBox.information(
                    self, "Already Started",
                    f"An activity for '{ticket_id}' is already running.\n"
                    "Please Stop it before starting a new one."
                )
                return
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to check activity:\n{e}")
            return

        now = datetime.now()
        try:
            append_activity_start(EXCEL_PATH, ticket_id, now)
        except PermissionError as e:
            QMessageBox.critical(self, "Cannot save", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to start activity:\n{e}")
            return

        self.update_status(f"Started activity for {ticket_id} at {now.strftime('%Y-%m-%d %H:%M:%S')}")
        if self.stack.currentWidget() is self.page_activity:
            self.load_activity_table()

    def on_stop_activity(self):
        ticket_id = self.current_ticket_from_selection_or_edit()
        if not ticket_id:
            QMessageBox.warning(self, "No Ticket ID", "Select or open a Ticket ID before stopping.")
            return

        now = datetime.now()
        try:
            updated = set_activity_end(EXCEL_PATH, ticket_id, now)
        except PermissionError as e:
            QMessageBox.critical(self, "Cannot save", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to stop activity:\n{e}")
            return

        if not updated:
            QMessageBox.information(
                self, "No open activity",
                f"No open activity found for '{ticket_id}'.\nStart one first."
            )
            return

        self.update_status(f"Stopped activity for {ticket_id} at {now.strftime('%Y-%m-%d %H:%M:%S')}")
        if self.stack.currentWidget() is self.page_activity:
            self.load_activity_table()

    # -------- Export to OneDrive --------

    def on_export_to_onedrive(self):
        onedrive_dir = find_onedrive_dir()
        default_dir = str(onedrive_dir)
        dest_dir = QFileDialog.getExistingDirectory(self, "Select OneDrive folder to export", default_dir)
        if not dest_dir:
            return
        dest_path = Path(dest_dir) / FILE_NAME

        if dest_path.exists():
            resp = QMessageBox.question(
                self, "Overwrite file?",
                f"'{dest_path}' already exists.\nDo you want to overwrite it?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if resp != QMessageBox.Yes:
                return

        try:
            # Atomic copy to destination
            tmp = dest_path.with_suffix(dest_path.suffix + f".tmp.{os.getpid()}")
            shutil.copy2(EXCEL_PATH, tmp)
            os.replace(tmp, dest_path)
        except Exception as e:
            QMessageBox.critical(self, "Export failed", f"Could not export to OneDrive:\n{e}")
            return

        self.update_status(f"Exported to {dest_path}")

    # -------- Common --------

    def open_excel(self):
        path = str(EXCEL_PATH)
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                import subprocess
                subprocess.run(["open", path], check=False)
            else:
                import subprocess
                subprocess.run(["xdg-open", path], check=False)
        except Exception as e:
            QMessageBox.critical(self, "Open failed", f"Could not open the file:\n{e}")

    def update_status(self, text: str):
        self.status.showMessage(text)


def main():
    # Windows: set AppUserModelID for proper taskbar icon/grouping
    if sys.platform == "win32":
        try:
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("Incident.Tracker.1.1")
        except Exception:
            pass

    ensure_workbook_and_sheet(EXCEL_PATH)

    app = QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)

    ico = load_window_icon()
    app.setWindowIcon(ico)

    w = MainWindow()
    w.setWindowIcon(ico)  # ensure window picks the same icon
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
