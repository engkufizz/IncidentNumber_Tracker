import os
import sys
import ctypes
from datetime import date, datetime
from pathlib import Path
from typing import List, Tuple

from PySide6.QtCore import Qt, QDate, QPoint
from PySide6.QtGui import QAction, QIcon, QPainter, QPixmap, QColor, QKeySequence
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QDateEdit, QTextEdit, QTableWidget, QTableWidgetItem,
    QAbstractItemView, QHeaderView, QMessageBox, QMenu, QStatusBar, QFrame, QLineEdit
)

# Excel
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("This tool requires 'openpyxl'. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

APP_TITLE = "Incident Number Tracker"
FILE_NAME = "incident_numbers.xlsx"
SHEET_NAME = "INCIDENTS"
HEADERS = ("Created On", "Ticket ID", "Description")
DATE_NUMBER_FORMAT = "yyyy-mm-dd"

# UI colours
ACCENT = "#C62828"       # red banner to differentiate
BG_LIGHT = "#FDF7F7"
ROW_ALT = "#FCEAEA"
TEXT_PRIMARY = "#2D1414"


def is_frozen() -> bool:
    return getattr(sys, "frozen", False)


def app_dir() -> Path:
    if is_frozen():
        return Path(sys.executable).resolve().parent
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd()


EXCEL_PATH = app_dir() / FILE_NAME


def ensure_workbook_and_sheet(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb["Sheet"])
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(title=SHEET_NAME)
        ws["A1"] = HEADERS[0]
        ws["B1"] = HEADERS[1]
        ws["C1"] = HEADERS[2]
    else:
        ws = wb[SHEET_NAME]
        if ws["A1"].value is None and ws["B1"].value is None and ws["C1"].value is None:
            ws["A1"] = HEADERS[0]
            ws["B1"] = HEADERS[1]
            ws["C1"] = HEADERS[2]
    wb.save(path)
    return wb


def normalize_description(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    parts = [line.strip() for line in text.split("\n")]
    parts = [p for p in parts if p]
    return ", ".join(parts)


def append_row(path: Path, d: date, ticket_id: str, desc: str):
    wb = ensure_workbook_and_sheet(path)
    ws = wb[SHEET_NAME]
    ws.append([d, ticket_id, desc])
    last_row = ws.max_row
    ws.cell(row=last_row, column=1).number_format = DATE_NUMBER_FORMAT
    try:
        wb.save(path)
    except PermissionError:
        raise PermissionError(
            f"Cannot save the Excel file.\n\nFile may be open or folder not writable:\n{path}\n\n"
            "Close the file if open, or move the EXE and Excel to a writable folder (e.g., Desktop/Documents)."
        )


def read_rows(path: Path) -> List[Tuple[str, str, str]]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[SHEET_NAME]
    rows: List[Tuple[str, str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        created_on, ticket_id, description = (row + (None, None, None))[:3]
        if created_on is None and ticket_id is None and description is None:
            continue
        if isinstance(created_on, (datetime, date)):
            dstr = created_on.strftime("%Y-%m-%d")
        elif created_on:
            dstr = str(created_on)
        else:
            dstr = ""
        rows.append((dstr, str(ticket_id or ""), description or ""))
    return rows


def next_default_ticket_for_date(path: Path, d: date) -> str:
    """
    Compute next 'THyymmddNN' for the given date by scanning existing Ticket IDs.
    Only considers IDs that start with the THyymmdd prefix.
    """
    yy = d.year % 100
    prefix = f"TH{yy:02d}{d.month:02d}{d.day:02d}"
    max_seq = 0
    try:
        if path.exists():
            wb = load_workbook(path, data_only=True)
            if SHEET_NAME in wb.sheetnames:
                ws = wb[SHEET_NAME]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    ticket_id = (row[1] or "").strip()
                    if ticket_id.startswith(prefix):
                        tail = ticket_id[len(prefix):]
                        if tail.isdigit():
                            max_seq = max(max_seq, int(tail))
    except Exception:
        pass
    return f"{prefix}{max_seq + 1:02d}"


def emoji_icon(emoji: str, size: int = 128,
               bg=QColor(198, 40, 40), fg=QColor(255, 255, 255)) -> QIcon:
    pm = QPixmap(size, size)
    pm.fill(Qt.transparent)
    painter = QPainter(pm)
    painter.setRenderHint(QPainter.Antialiasing, True)
    painter.setBrush(bg)
    painter.setPen(Qt.NoPen)
    painter.drawEllipse(0, 0, size, size)
    painter.setPen(fg)
    painter.drawText(pm.rect(), Qt.AlignCenter, emoji)
    painter.end()
    return QIcon(pm)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.setMinimumSize(900, 650)
        self.setWindowIcon(emoji_icon("🚨"))  # distinct logo

        ensure_workbook_and_sheet(EXCEL_PATH)

        self._user_edited_ticket = False  # track if user has typed in Ticket ID

        central = QWidget()
        root = QVBoxLayout(central)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        # Banner
        banner = QFrame()
        banner.setStyleSheet(f"background:{ACCENT}; border-radius:6px;")
        bl = QVBoxLayout(banner)
        title = QLabel(APP_TITLE)
        subtitle = QLabel("Track incident tickets quickly (with smart default IDs)")
        title.setStyleSheet("color:white; font-size:18px; font-weight:600;")
        subtitle.setStyleSheet("color:#FFE5E5; font-size:12px;")
        bl.addWidget(title)
        bl.addWidget(subtitle)
        root.addWidget(banner)

        # Top row: date + ticket id
        top = QHBoxLayout()

        lbl_date = QLabel("Created On:")
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.dateChanged.connect(self.on_date_changed)
        btn_today = QPushButton("Today")
        btn_today.clicked.connect(self.set_today)

        top.addWidget(lbl_date)
        top.addWidget(self.date_edit)
        top.addWidget(btn_today)
        top.addSpacing(20)

        lbl_ticket = QLabel("Ticket ID:")
        self.ticket_edit = QLineEdit()
        self.ticket_edit.setPlaceholderText("e.g. INC2453903 or leave default THyymmddNN")
        self.ticket_edit.textEdited.connect(self.on_ticket_edited)
        btn_fill_default = QPushButton("Fill Default")
        btn_fill_default.clicked.connect(self.fill_default_ticket)

        top.addWidget(lbl_ticket)
        top.addWidget(self.ticket_edit)
        top.addWidget(btn_fill_default)
        top.addStretch(1)
        root.addLayout(top)

        # Description
        desc_frame = QVBoxLayout()
        lbl_desc = QLabel("Description (multi-line allowed)")
        self.desc_text = QTextEdit()
        self.desc_text.setPlaceholderText("Enter incident description; multiple lines will be joined with commas")
        self.desc_text.textChanged.connect(self.update_preview)
        desc_frame.addWidget(lbl_desc)
        desc_frame.addWidget(self.desc_text)
        root.addLayout(desc_frame)

        # Preview
        lbl_prev = QLabel("Preview (single line):")
        self.preview = QLabel("(nothing yet)")
        self.preview.setStyleSheet("background:white; border:1px solid #E6BFBF; padding:8px;")
        root.addWidget(lbl_prev)
        root.addWidget(self.preview)

        # Buttons
        btn_row = QHBoxLayout()
        self.btn_add = QPushButton("Add")
        self.btn_add.clicked.connect(self.on_add)
        self.btn_clear = QPushButton("Clear")
        self.btn_clear.clicked.connect(self.on_clear)
        self.btn_open = QPushButton("Open Excel")
        self.btn_open.clicked.connect(self.open_excel)
        self.btn_refresh = QPushButton("Refresh List")
        self.btn_refresh.clicked.connect(self.load_table)

        tip = QLabel("Shortcuts: Ctrl+Enter Add • Ctrl+L Clear • Ctrl+T Today • Ctrl+O Open • F5 Refresh")
        tip.setStyleSheet("color:#8A5555;")

        btn_row.addWidget(self.btn_add)
        btn_row.addWidget(self.btn_clear)
        btn_row.addWidget(self.btn_open)
        btn_row.addWidget(self.btn_refresh)
        btn_row.addStretch(1)
        btn_row.addWidget(tip)
        root.addLayout(btn_row)

        # Table (3 columns)
        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(HEADERS)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.on_table_context_menu)
        root.addWidget(self.table)

        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.update_status(f"File: {EXCEL_PATH}")

        self.setCentralWidget(central)

        # Shortcuts
        self._add_shortcuts()

        # Initial fill and load
        self.fill_default_ticket()
        self.load_table()
        self.update_preview()

        # Styling
        self._apply_styles()

    def _apply_styles(self):
        self.setStyleSheet(f"""
            QWidget {{ background: {BG_LIGHT}; color: {TEXT_PRIMARY}; }}
            QTextEdit {{ background: white; border: 1px solid #E6BFBF; }}
            QTableWidget {{ background: white; alternate-background-color: {ROW_ALT}; }}
            QPushButton {{ padding: 6px 10px; }}
        """)

    def _add_shortcuts(self):
        def add_seq(seq, handler):
            act = QAction(self)
            act.setShortcut(QKeySequence(seq))
            act.triggered.connect(handler)
            self.addAction(act)

        add_seq("Ctrl+Return", self.on_add)
        add_seq("Ctrl+Enter", self.on_add)
        add_seq("Ctrl+L", self.on_clear)
        add_seq("Ctrl+T", self.set_today)
        add_seq("Ctrl+O", self.open_excel)
        add_seq("F5", self.load_table)

    def set_today(self):
        self.date_edit.setDate(QDate.currentDate())

    def on_date_changed(self):
        # Only auto-update if the user hasn't started typing a manual ID
        if not self._user_edited_ticket:
            self.fill_default_ticket()

    def on_ticket_edited(self):
        self._user_edited_ticket = True

    def fill_default_ticket(self):
        dqt = self.date_edit.date()
        d_py = date(dqt.year(), dqt.month(), dqt.day())
        suggested = next_default_ticket_for_date(EXCEL_PATH, d_py)
        self.ticket_edit.setText(suggested)
        self._user_edited_ticket = False  # treat the current value as app-suggested

    def on_clear(self):
        self.desc_text.clear()
        self.update_preview()
        # Keep date and ticket; often you’ll add multiple entries

    def on_add(self):
        # Date
        dqt = self.date_edit.date()
        d_py = date(dqt.year(), dqt.month(), dqt.day())

        # Ticket ID
        ticket_id = self.ticket_edit.text().strip()
        if not ticket_id:
            ticket_id = next_default_ticket_for_date(EXCEL_PATH, d_py)

        desc = normalize_description(self.desc_text.toPlainText())
        if not desc:
            QMessageBox.critical(self, "Missing description", "Please enter the Description.")
            return

        try:
            append_row(EXCEL_PATH, d_py, ticket_id, desc)
        except PermissionError as e:
            QMessageBox.critical(self, "Cannot save", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add entry:\n{e}")
            return

        self.update_status(f"Added: {d_py.strftime('%Y-%m-%d')} • {ticket_id}")

        # Prepare for next entry:
        # 1) Clear description
        self.on_clear()
        # 2) Always set Ticket ID to the next default to avoid reuse
        self.fill_default_ticket()
        # 3) Refresh table
        self.load_table()

    def update_preview(self):
        text = self.desc_text.toPlainText().strip()
        combined = normalize_description(text)
        self.preview.setText(combined or "(nothing yet)")

    def load_table(self):
        try:
            ensure_workbook_and_sheet(EXCEL_PATH)
            rows = read_rows(EXCEL_PATH)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read Excel:\n{e}")
            return

        self.table.setRowCount(0)
        for dstr, ticket, desc in rows:
            r = self.table.rowCount()
            self.table.insertRow(r)
            self.table.setItem(r, 0, QTableWidgetItem(dstr))
            self.table.setItem(r, 1, QTableWidgetItem(ticket))
            self.table.setItem(r, 2, QTableWidgetItem(desc))

        self.update_status(f"File: {EXCEL_PATH} • records: {len(rows)}")

    def on_table_context_menu(self, pos: QPoint):
        idx = self.table.indexAt(pos)
        if not idx.isValid():
            return
        self.table.selectRow(idx.row())
        menu = QMenu(self)
        act_copy = QAction("Copy row", self)
        act_copy.triggered.connect(self.copy_selected_row)
        menu.addAction(act_copy)
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def copy_selected_row(self):
        row = self.table.currentRow()
        if row < 0:
            return
        d = self.table.item(row, 0).text() if self.table.item(row, 0) else ""
        ticket = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
        desc = self.table.item(row, 2).text() if self.table.item(row, 2) else ""
        text = f"{d}\t{ticket}\t{desc}"
        QApplication.clipboard().setText(text)
        self.update_status("Row copied to clipboard.")

    def open_excel(self):
        path = str(EXCEL_PATH)
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                import subprocess
                subprocess.run(["open", path], check=False)
            else:
                import subprocess
                subprocess.run(["xdg-open", path], check=False)
        except Exception as e:
            QMessageBox.critical(self, "Open failed", f"Could not open the file:\n{e}")

    def update_status(self, text: str):
        self.status.showMessage(text)


def main():
    # Windows: set AppUserModelID for proper taskbar icon/grouping
    if sys.platform == "win32":
        try:
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("IncidentTracker.Maxis.1.0")
        except Exception:
            pass

    ensure_workbook_and_sheet(EXCEL_PATH)

    app = QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)
    app.setWindowIcon(emoji_icon("🚨"))

    w = MainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
